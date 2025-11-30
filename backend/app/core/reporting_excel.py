"""
Excel report generation using openpyxl.
Creates multi-sheet workbooks with charts and professional formatting.
"""

from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List
from pathlib import Path
from datetime import datetime
import json


def generate_excel_report(
    project_id: str,
    project_data: Dict[str, Any],
    segmentation_data: Dict[str, Any],
    analysis_data: Dict[str, Any],
    output_path: str
) -> str:
    """
    Generate comprehensive Excel report.
    
    Args:
        project_id: Project identifier
        project_data: Project metadata
        segmentation_data: Segmentation results
        analysis_data: AI analysis results
        output_path: Path to save Excel file
    
    Returns:
        Path to generated Excel file
    """
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create sheets
    _create_overview_sheet(wb, project_id, project_data, segmentation_data, analysis_data)
    _create_frames_sheet(wb, segmentation_data)
    _create_objects_sheet(wb, segmentation_data)
    _create_ai_insights_sheet(wb, analysis_data)
    
    # Save workbook
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    return output_path


def _create_overview_sheet(wb, project_id, project_data, segmentation_data, analysis_data):
    """Create overview sheet with summary statistics and charts."""
    ws = wb.create_sheet("Overview", 0)
    
    # Title
    ws['A1'] = "Ciousten - Video Analysis Report"
    ws['A1'].font = Font(size=16, bold=True, color="1F4E78")
    ws.merge_cells('A1:D1')
    
    # Project info
    ws['A3'] = "Project Information"
    ws['A3'].font = Font(size=12, bold=True)
    
    ws['A4'] = "Project ID:"
    ws['B4'] = project_id
    ws['A5'] = "Video File:"
    ws['B5'] = project_data.get('video_filename', 'N/A')
    ws['A6'] = "Generated:"
    ws['B6'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['A7'] = "Created by:"
    ws['B7'] = "Aditya Shenvi @2025 (www.adityacuz.dev)"
    
    # Segmentation stats
    ws['A9'] = "Segmentation Statistics"
    ws['A9'].font = Font(size=12, bold=True)
    
    stats = segmentation_data.get('stats', {})
    ws['A10'] = "Total Frames:"
    ws['B10'] = stats.get('total_frames', 0)
    ws['A11'] = "Total Objects:"
    ws['B11'] = stats.get('total_objects', 0)
    ws['A12'] = "Unique Objects:"
    ws['B12'] = stats.get('unique_objects', 'N/A')
    ws['A13'] = "Avg Objects/Frame:"
    ws['B13'] = round(stats.get('avg_objects_per_frame', 0), 2)
    ws['A14'] = "Processing Time:"
    ws['B14'] = f"{stats.get('processing_time_seconds', 0):.1f}s"
    
    # Objects per class
    ws['A15'] = "Objects by Class"
    ws['A15'].font = Font(size=12, bold=True)
    
    objects_per_class = stats.get('objects_per_class', {})
    row = 16
    ws['A16'] = "Class"
    ws['B16'] = "Count"
    ws['A16'].font = Font(bold=True)
    ws['B16'].font = Font(bold=True)
    
    for class_name, count in objects_per_class.items():
        row += 1
        ws[f'A{row}'] = class_name
        ws[f'B{row}'] = count
    
    # Add pie chart for class distribution
    if len(objects_per_class) > 0:
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=17, max_row=row)
        data = Reference(ws, min_col=2, min_row=16, max_row=row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Object Class Distribution"
        ws.add_chart(pie, "D9")
    
    # AI Summary
    ws['A30'] = "AI Analysis Summary"
    ws['A30'].font = Font(size=12, bold=True)
    ws['A31'] = analysis_data.get('summary', 'No analysis available')
    ws.merge_cells('A31:F35')
    ws['A31'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30


def _create_frames_sheet(wb, segmentation_data):
    """Create frames sheet with per-frame statistics."""
    ws = wb.create_sheet("Frames")
    
    # Headers
    headers = ["Frame Index", "Timestamp (s)", "Object Count", "Classes Detected"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Data
    frames = segmentation_data.get('frames', [])
    for row, frame in enumerate(frames, 2):
        ws.cell(row, 1, frame.get('frame_index', 0))
        ws.cell(row, 2, round(frame.get('timestamp', 0), 2))
        ws.cell(row, 3, len(frame.get('objects', [])))
        
        # Get unique classes in this frame
        classes = set(obj.get('class_name', 'unknown') for obj in frame.get('objects', []))
        ws.cell(row, 4, ', '.join(sorted(classes)))
    
    # Add line chart for object count over time
    if len(frames) > 1:
        chart = LineChart()
        chart.title = "Objects Detected Over Time"
        chart.y_axis.title = "Object Count"
        chart.x_axis.title = "Frame Index"
        
        data = Reference(ws, min_col=3, min_row=1, max_row=len(frames) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(frames) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "F2")
    
    # Adjust column widths
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20


def _create_objects_sheet(wb, segmentation_data):
    """Create objects sheet with all detected objects."""
    ws = wb.create_sheet("Objects")
    
    # Headers
    headers = ["Frame", "Object ID", "Class", "Confidence", "BBox (x1,y1,x2,y2)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Data
    row = 2
    frames = segmentation_data.get('frames', [])
    for frame in frames:
        frame_idx = frame.get('frame_index', 0)
        for obj in frame.get('objects', []):
            ws.cell(row, 1, frame_idx)
            ws.cell(row, 2, obj.get('id', 0))
            ws.cell(row, 3, obj.get('class_name', 'unknown'))
            ws.cell(row, 4, round(obj.get('confidence', 0), 3))
            
            bbox = obj.get('bbox', [0, 0, 0, 0])
            bbox_str = f"({bbox[0]:.0f},{bbox[1]:.0f},{bbox[2]:.0f},{bbox[3]:.0f})"
            ws.cell(row, 5, bbox_str)
            
            row += 1
    
    # Adjust column widths
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _create_ai_insights_sheet(wb, analysis_data):
    """Create AI insights sheet with analysis results."""
    ws = wb.create_sheet("AI_Insights")
    
    # Title
    ws['A1'] = "AI-Generated Insights"
    ws['A1'].font = Font(size=14, bold=True, color="1F4E78")
    
    # Summary
    ws['A3'] = "Summary"
    ws['A3'].font = Font(size=12, bold=True)
    ws['A4'] = analysis_data.get('summary', 'No summary available')
    ws.merge_cells('A4:F8')
    ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Key Findings
    ws['A10'] = "Key Findings"
    ws['A10'].font = Font(size=12, bold=True)
    
    findings = analysis_data.get('key_findings', [])
    for i, finding in enumerate(findings, 11):
        ws[f'A{i}'] = f"• {finding}"
        ws[f'A{i}'].alignment = Alignment(wrap_text=True)
    
    # Anomalies
    anomaly_row = 11 + len(findings) + 2
    ws[f'A{anomaly_row}'] = "Anomalies Detected"
    ws[f'A{anomaly_row}'].font = Font(size=12, bold=True)
    
    anomalies = analysis_data.get('anomalies', [])
    for i, anomaly in enumerate(anomalies):
        row = anomaly_row + i + 1
        ws[f'A{row}'] = f"• {anomaly}"
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
    
    # KPIs
    kpi_row = anomaly_row + len(anomalies) + 3
    ws[f'A{kpi_row}'] = "Key Performance Indicators"
    ws[f'A{kpi_row}'].font = Font(size=12, bold=True)
    
    ws[f'A{kpi_row + 1}'] = "Metric"
    ws[f'B{kpi_row + 1}'] = "Value"
    ws[f'C{kpi_row + 1}'] = "Unit"
    
    kpis = analysis_data.get('kpis', [])
    for i, kpi in enumerate(kpis):
        row = kpi_row + i + 2
        ws[f'A{row}'] = kpi.get('name', '')
        ws[f'B{row}'] = kpi.get('value', 0)
        ws[f'C{row}'] = kpi.get('unit', '')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
